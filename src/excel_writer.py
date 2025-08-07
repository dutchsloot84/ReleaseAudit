"""Excel report generation utilities."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)


def _merge_on_key(old_df: pd.DataFrame, new_df: pd.DataFrame, key: Optional[str], mode: str) -> pd.DataFrame:
    """Merge ``new_df`` into ``old_df`` using ``mode`` on ``key`` column."""
    if mode == "replace" or key is None or key not in new_df.columns or key not in old_df.columns:
        return new_df

    if mode == "append":
        combined = pd.concat([old_df, new_df], ignore_index=True)
        return combined.drop_duplicates(subset=key, keep="last")

    if mode == "upsert":
        old = old_df.set_index(key)
        new = new_df.set_index(key)
        old.update(new)
        combined = pd.concat([old, new[~new.index.isin(old.index)]])
        return combined.reset_index()

    raise ValueError(f"Unknown merge mode: {mode}")


def _as_table(ws, name_hint: str) -> None:
    """Convert the worksheet data range into an Excel table."""
    if ws.max_row == 0 or ws.max_column == 0:
        return
    ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table_name = "T_" + "".join(ch if ch.isalnum() else "_" for ch in name_hint)[:25]
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)


def _autofit(ws) -> None:
    """Auto-size columns with a maximum width."""
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        width = min(length + 2, 60)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def _wrap_columns(ws, cols: Iterable[str]) -> None:
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    indexes = [header.index(c) for c in cols if c in header]
    for row in ws.iter_rows(min_row=2):
        for idx in indexes:
            row[idx].alignment = row[idx].alignment.copy(wrap_text=True)


def _hyperlink_column(ws, url_col: str, text_col: Optional[str] = None) -> None:
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        url_idx = header.index(url_col)
    except ValueError:
        return
    text_idx = header.index(text_col) if text_col in header else None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        url_cell = row[url_idx]
        url = url_cell.value
        if not url:
            continue
        display = url
        if text_idx is not None:
            text_value = row[text_idx].value or ""
            display = str(text_value)
            if text_col and "commit" in text_col.lower():
                display = display[:7]
        url_cell.value = display
        url_cell.hyperlink = url
        url_cell.style = "Hyperlink"


def _freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


def write_report(
    output_file: str,
    data_frames: Dict[str, pd.DataFrame],
    hyperlink_specs: Optional[Dict[str, List[Dict]]],
    update_existing: bool = False,
    merge_mode: str = "upsert",
    primary_keys: Optional[Dict[str, Optional[str]]] = None,
) -> None:
    """Create or update an Excel report with rich formatting and hyperlinks."""

    output_path = Path(output_file)
    final_frames: Dict[str, pd.DataFrame] = {}

    if update_existing and output_path.exists():
        excel = pd.ExcelFile(output_path)
        for sheet, new_df in data_frames.items():
            if sheet in excel.sheet_names:
                old_df = excel.parse(sheet)
                key = primary_keys.get(sheet) if primary_keys else None
                merged = _merge_on_key(old_df, new_df, key, merge_mode)
            else:
                merged = new_df
            final_frames[sheet] = merged
    else:
        final_frames = data_frames

    wb = Workbook()
    wb.remove(wb.active)

    for sheet, df in final_frames.items():
        ws = wb.create_sheet(title=sheet[:31])
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        _as_table(ws, sheet)
        _freeze_header(ws)
        _autofit(ws)
        wrap_cols = [c for c in ("Summary", "Message") if c in df.columns]
        if wrap_cols:
            _wrap_columns(ws, wrap_cols)
        if hyperlink_specs and sheet in hyperlink_specs:
            for spec in hyperlink_specs[sheet]:
                _hyperlink_column(ws, spec.get("url_col"), spec.get("text_col"))

    wb.save(output_file)
